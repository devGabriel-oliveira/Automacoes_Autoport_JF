"""Automações Gestão de Frotas - Autoport — ARQUIVO ÚNICO.

Tudo em um só arquivo de propósito: assim não existe pasta `core/` para
faltar no repositório, e o ModuleNotFoundError não pode acontecer.

Rodar:  streamlit run app.py
"""

from __future__ import annotations

import csv
import datetime as dt
import random
import re
import time
import urllib.parse
import webbrowser
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

# ======================================================================
# MODULO: telefone
# ======================================================================
"""Normalização de telefones brasileiros para o formato aceito pelo WhatsApp.

Formato de saída: 55 + DDD (2) + número (9) = 13 dígitos.
"""



# Valores que a planilha usa para "não tem telefone"
VAZIOS = {"", "-", "#N/A", "#N/D", "NAN", "NONE", "NULL", "0"}

# DDDs válidos no Brasil
DDDS_VALIDOS = {
    11, 12, 13, 14, 15, 16, 17, 18, 19,
    21, 22, 24, 27, 28,
    31, 32, 33, 34, 35, 37, 38,
    41, 42, 43, 44, 45, 46, 47, 48, 49,
    51, 53, 54, 55,
    61, 62, 63, 64, 65, 66, 67, 68, 69,
    71, 73, 74, 75, 77, 79,
    81, 82, 83, 84, 85, 86, 87, 88, 89,
    91, 92, 93, 94, 95, 96, 97, 98, 99,
}

# Motivos de descarte / alerta
OK = "ok"
INFERIDO = "inferido"          # tivemos que adicionar o 9 — pode ser fixo
SEM_CADASTRO = "sem_cadastro"
INVALIDO = "invalido"


def normalizar_telefone(valor) -> str | None:
    """Devolve o telefone em formato WhatsApp (13 dígitos) ou None."""
    numero, _ = normalizar_com_motivo(valor)
    return numero


def normalizar_com_motivo(valor) -> tuple[str | None, str]:
    """Igual a normalizar_telefone, mas devolve também o motivo.

    Retorna (numero, motivo). Motivo é um dos: OK, INFERIDO,
    SEM_CADASTRO, INVALIDO.
    """
    if valor is None:
        return None, SEM_CADASTRO

    texto = str(valor).strip()
    if texto.upper() in VAZIOS:
        return None, SEM_CADASTRO

    digitos = re.sub(r"\D", "", texto)
    if not digitos:
        return None, SEM_CADASTRO

    # Já veio com DDI
    if len(digitos) == 13 and digitos.startswith("55"):
        return (digitos, OK) if _ddd_valido(digitos[2:4]) else (None, INVALIDO)

    # DDI + DDD + 8 dígitos (falta o 9)
    if len(digitos) == 12 and digitos.startswith("55"):
        ddd, resto = digitos[2:4], digitos[4:]
        if not _ddd_valido(ddd):
            return None, INVALIDO
        return f"55{ddd}9{resto}", INFERIDO

    # DDD + 9 dígitos
    if len(digitos) == 11:
        ddd = digitos[:2]
        if not _ddd_valido(ddd):
            return None, INVALIDO
        return f"55{digitos}", OK

    # DDD + 8 dígitos — precisa inferir o 9
    if len(digitos) == 10:
        ddd, resto = digitos[:2], digitos[2:]
        if not _ddd_valido(ddd):
            return None, INVALIDO
        return f"55{ddd}9{resto}", INFERIDO

    return None, INVALIDO


def _ddd_valido(ddd: str) -> bool:
    try:
        return int(ddd) in DDDS_VALIDOS
    except (TypeError, ValueError):
        return False


def formatar_exibicao(numero: str | None) -> str:
    """Formata 5531982634233 como +55 (31) 98263-4233 para mostrar na tela."""
    if not numero or len(numero) != 13:
        return "—"
    return f"+55 ({numero[2:4]}) {numero[4:9]}-{numero[9:]}"


# ======================================================================
# MODULO: leitura
# ======================================================================
"""Carga e limpeza da planilha GERAL-NOVO_MAPA_PORTO.xlsx."""





STATUS_PENDENTES = {"Vencido", "A Vencer em 30 dias"}


@dataclass(frozen=True)
class Modulo:
    """Configuração de leitura de uma aba."""

    chave: str
    rotulo: str
    aba: str          # nome exato na planilha (atenção aos espaços!)
    col_item: str     # o que está vencendo (placa, nome, frota...)
    col_validade: str
    col_status: str | None
    col_telefone: str
    col_email: str | None
    col_empresa: str | None
    rotulo_item: str
    fonte: str = "porto"


MODULOS: dict[str, Modulo] = {
    "cavalo": Modulo(
        chave="cavalo",
        rotulo="Cavalo",
        aba="Cavalo",
        col_item="CAVALO",
        col_validade="VALIDADE",
        col_status="STATUS",
        col_telefone="Telefone",
        col_email="E-mail",
        col_empresa="EMPRESA",
        rotulo_item="Placa",
    ),
    "motorista": Modulo(
        chave="motorista",
        rotulo="Motorista",
        aba="Motorista",
        col_item="MOTORISTA",
        col_validade="VALIDADE",
        col_status="Status",
        col_telefone="Telefone",
        col_email="E-mail",
        col_empresa="Empresa",
        rotulo_item="Motorista",
    ),
    "carreta": Modulo(
        chave="carreta",
        rotulo="Carreta",
        aba="CARRETA",
        col_item="PRANCHA/CEGONHA",
        col_validade="VALIDADE",
        col_status="STATUS",
        col_telefone="TELEFONE",
        col_email="E-MAIL",
        col_empresa="EMPRESA",
        rotulo_item="Placa",
    ),
    "empresas": Modulo(
        chave="empresas",
        rotulo="EMPRESAS-VENCIMENTOS",
        aba="EMPRESAS-VENCIMENTOS",
        col_item="EMPRESA",
        col_validade="VENCIMENTO PORTO",
        col_status=None,           # esta aba não tem coluna de status usável
        col_telefone="TELEFONE",
        col_email="E-MAIL",
        col_empresa="EMPRESA",
        rotulo_item="Empresa",
    ),
}

DIAS_ALERTA = 30


@dataclass
class Registro:
    """Uma linha pendente da planilha."""

    item: str
    validade: dt.date | None
    status: str
    empresa: str
    telefone: str | None
    telefone_motivo: str
    telefone_exibicao: str
    email: str
    linha_planilha: int
    documento: str = ""   # qual documento venceu (abas com várias colunas)

    @property
    def enviavel(self) -> bool:
        return self.telefone is not None

    @property
    def descricao(self) -> str:
        """'RBO0J52 — CRONOTACÓGRAFO' ou só 'RBO0J52'."""
        return f"{self.item} — {self.documento}" if self.documento else self.item

    @property
    def validade_texto(self) -> str:
        return self.validade.strftime("%d/%m/%Y") if self.validade else "sem data"


@dataclass
class Contato:
    """Um destinatário e todos os itens que ele precisa regularizar."""

    telefone: str
    telefone_exibicao: str
    empresa: str
    registros: list[Registro] = field(default_factory=list)

    @property
    def tem_numero_inferido(self) -> bool:
        return any(r.telefone_motivo == INFERIDO for r in self.registros)

    @property
    def qtd(self) -> int:
        return len(self.registros)


def _normalizar_nomes(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    return df


def carregar_aba(caminho: Path | str, modulo: Modulo) -> pd.DataFrame:
    """Lê a aba tolerando espaços sobrando no nome."""
    xls = pd.ExcelFile(caminho)
    alvo = modulo.aba.strip().upper()
    nome_real = next(
        (n for n in xls.sheet_names if str(n).strip().upper() == alvo), None
    )
    if nome_real is None:
        disponiveis = ", ".join(repr(n) for n in xls.sheet_names)
        raise KeyError(
            f"Aba {modulo.aba!r} não encontrada. Abas disponíveis: {disponiveis}"
        )
    return _normalizar_nomes(pd.read_excel(xls, sheet_name=nome_real))


def _para_data(valor) -> dt.date | None:
    """Converte para date. Texto como 'TERMO' vira None sem quebrar."""
    # pd.NaT é instância de datetime, então precisa ser barrado ANTES
    # do isinstance abaixo — senão vira uma data que quebra no strftime.
    if valor is None or valor is pd.NaT:
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    try:
        convertido = pd.to_datetime(valor, dayfirst=True, errors="coerce")
    except (ValueError, TypeError):
        return None
    if pd.isna(convertido):
        return None
    data = convertido.date()
    # A planilha usa 01/01/1900 como marcador de "sem data"
    return None if data.year < 1990 else data


def _limpar_email(valor) -> str:
    """'Transportes Las Vegas <x@y.com>' -> 'x@y.com'."""
    if valor is None or pd.isna(valor):
        return ""
    texto = str(valor).strip()
    if texto.upper() in {"#N/A", "-", "NAN"}:
        return ""
    if "<" in texto and ">" in texto:
        return texto[texto.index("<") + 1 : texto.index(">")].strip()
    return texto


def _status_por_data(validade: dt.date | None, hoje: dt.date) -> str:
    if validade is None:
        return "Sem data"
    dias = (validade - hoje).days
    if dias < 0:
        return "Vencido"
    if dias <= DIAS_ALERTA:
        return "A Vencer em 30 dias"
    return "Não Vencida"


def ler_registros(
    caminho: Path | str, modulo: Modulo, hoje: dt.date | None = None
) -> list[Registro]:
    """Devolve apenas os registros pendentes (vencidos ou a vencer em 30 dias)."""
    hoje = hoje or dt.date.today()
    df = carregar_aba(caminho, modulo)

    faltando = [
        c
        for c in (modulo.col_item, modulo.col_validade, modulo.col_telefone)
        if c not in df.columns
    ]
    if faltando:
        raise KeyError(
            f"Aba {modulo.aba!r}: colunas ausentes {faltando}. "
            f"Encontradas: {list(df.columns)}"
        )

    registros: list[Registro] = []

    for posicao, linha in df.iterrows():
        item = linha.get(modulo.col_item)
        if item is None or pd.isna(item) or not str(item).strip():
            continue

        validade = _para_data(linha.get(modulo.col_validade))

        if modulo.col_status and modulo.col_status in df.columns:
            status = str(linha.get(modulo.col_status) or "").strip()
            if status not in STATUS_PENDENTES:
                continue
        else:
            status = _status_por_data(validade, hoje)
            if status not in STATUS_PENDENTES:
                continue

        telefone, motivo = normalizar_com_motivo(linha.get(modulo.col_telefone))

        empresa = ""
        if modulo.col_empresa and modulo.col_empresa in df.columns:
            bruto = linha.get(modulo.col_empresa)
            empresa = "" if pd.isna(bruto) else str(bruto).strip()

        email = ""
        if modulo.col_email and modulo.col_email in df.columns:
            email = _limpar_email(linha.get(modulo.col_email))

        registros.append(
            Registro(
                item=str(item).strip(),
                validade=validade,
                status=status,
                empresa=empresa or "(sem empresa)",
                telefone=telefone,
                telefone_motivo=motivo,
                telefone_exibicao=formatar_exibicao(telefone),
                email=email,
                linha_planilha=int(posicao) + 2,
            )
        )

    return registros


def agrupar_por_contato(registros: list[Registro]) -> list[Contato]:
    """Agrupa por telefone: 8 cavalos de uma empresa = 1 mensagem, não 8."""
    contatos: dict[str, Contato] = {}
    for reg in registros:
        if not reg.enviavel:
            continue
        assert reg.telefone is not None
        if reg.telefone not in contatos:
            contatos[reg.telefone] = Contato(
                telefone=reg.telefone,
                telefone_exibicao=reg.telefone_exibicao,
                empresa=reg.empresa,
            )
        contatos[reg.telefone].registros.append(reg)
    return sorted(contatos.values(), key=lambda c: c.empresa)


def pendencias_de_cadastro(registros: list[Registro]) -> list[Registro]:
    """Registros pendentes que não podem ser cobrados por falta de telefone."""
    return [r for r in registros if not r.enviavel]


def resumo(caminho: Path | str, hoje: dt.date | None = None) -> dict[str, dict]:
    """Contagens por módulo, para os cards da tela inicial."""
    saida: dict[str, dict] = {}
    for chave, modulo in MODULOS.items():
        try:
            registros = ler_registros(caminho, modulo, hoje)
        except (KeyError, ValueError) as erro:
            saida[chave] = {"erro": str(erro), "pendentes": 0, "contatos": 0,
                            "sem_telefone": 0, "vencidos": 0, "a_vencer": 0}
            continue
        saida[chave] = {
            "erro": None,
            "pendentes": len(registros),
            "vencidos": sum(1 for r in registros if r.status == "Vencido"),
            "a_vencer": sum(1 for r in registros if r.status == "A Vencer em 30 dias"),
            "sem_telefone": len(pendencias_de_cadastro(registros)),
            "contatos": len(agrupar_por_contato(registros)),
        }
    return saida


# ======================================================================
# MODULO: leitura_frotas
# ======================================================================
"""Leitura da planilha Mapa_Frotas.xlsx.

Diferença estrutural em relação ao mapa do porto: aqui cada LINHA tem VÁRIAS
colunas de documento, cada uma com sua própria data de validade. Uma linha pode
gerar vários alertas (ex.: o mesmo cavalo com Cronotacógrafo e CRLV vencidos).

O arquivo é grande (~35 MB) e tem uma aba com 1 milhão de linhas de #REF!,
então a leitura usa openpyxl em modo read_only — pandas carregaria tudo.
"""





DIAS_ALERTA = 30

# Datas usadas como marcador, não como vencimento real
ANO_MINIMO = 1990
ANO_MAXIMO = 2500


@dataclass(frozen=True)
class ModuloFrotas:
    """Configuração de uma aba do Mapa_Frotas (índices de coluna base 0)."""

    chave: str
    rotulo: str
    aba: str
    linha_cabecalho: int
    col_item: int
    col_telefone: int
    col_empresa: int
    col_contato: int | None
    col_frota: int | None
    documentos: dict[int, str]           # coluna -> nome do documento
    col_crlv_ano: int | None = None      # coluna com ANO (não data)
    rotulo_item: str = "Item"
    fonte: str = "frotas"
    extras: dict[int, str] = field(default_factory=dict)  # colunas informativas


MODULOS_FROTAS: dict[str, ModuloFrotas] = {
    "cavalo_carreta": ModuloFrotas(
        chave="cavalo_carreta",
        rotulo="Cavalo-Carreta",
        aba="CAVALO-CARRETA",
        linha_cabecalho=2,
        col_item=4,          # E - PLACA CAVALO
        col_telefone=2,      # C - Telefone
        col_empresa=1,       # B - empresa (cabeçalho em branco na planilha)
        col_contato=3,       # D - CONTATO
        col_frota=5,         # F - FROTA
        documentos={
            6: "Licença Estadual SP (Cavalo) Cegonha",
            7: "Licença Estadual SP (Cavalo) Prancha",
            8: "Cronotacógrafo",
            9: "Índice de Fumaça",
        },
        col_crlv_ano=10,     # K - CRLV CAVALO (ano, não data)
        rotulo_item="Placa",
    ),
    "motoristas_agregados": ModuloFrotas(
        chave="motoristas_agregados",
        rotulo="Motoristas Agregados",
        aba="MOTORISTAS AGREGADOS",
        linha_cabecalho=2,
        col_item=3,          # D - MOTORISTA
        col_telefone=2,      # C - TELEFONE
        col_empresa=1,       # B - EMPRESA
        col_contato=None,
        col_frota=None,
        documentos={
            5: "Direção Defensiva",
            6: "ASO",
            7: "CNH",
            8: "Toxicológico",
        },
        rotulo_item="Motorista",
    ),
    "contrato_aluguel": ModuloFrotas(
        chave="contrato_aluguel",
        rotulo="Venc. Contrato - Aluguel",
        aba="VENC. CONTRATO - ALUGUEL",
        linha_cabecalho=1,
        col_item=0,          # A - LOCATÁRIO
        col_telefone=1,      # B - Telefone
        col_empresa=0,       # A - LOCATÁRIO
        col_contato=None,
        col_frota=None,
        documentos={3: "Locação da credencial"},
        rotulo_item="Locatário",
        extras={2: "Credencial locada", 4: "Locador"},
    ),
}


def _para_data_frotas(valor) -> dt.date | None:
    if valor is None:
        return None
    if isinstance(valor, dt.datetime):
        data = valor.date()
    elif isinstance(valor, dt.date):
        data = valor
    else:
        return None
    if data.year < ANO_MINIMO or data.year > ANO_MAXIMO:
        return None  # 01/01/1900 e 01/01/2999 são marcadores, não vencimento
    return data


def _situacao(data: dt.date | None, hoje: dt.date) -> str | None:
    """None = não é alerta."""
    if data is None:
        return None
    dias = (data - hoje).days
    if dias < 0:
        return "Vencido"
    if dias <= DIAS_ALERTA:
        return "A Vencer em 30 dias"
    return None


def _texto(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    return "" if texto in {"#N/A", "#REF!", "#VALUE!", "-", "nan", "None"} else texto


def ler_registros_frotas(
    caminho: Path | str, modulo: ModuloFrotas, hoje: dt.date | None = None
) -> list[Registro]:
    """Um Registro por DOCUMENTO em alerta (não por linha)."""
    hoje = hoje or dt.date.today()

    livro = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    try:
        alvo = modulo.aba.strip().upper()
        nome_real = next(
            (n for n in livro.sheetnames if str(n).strip().upper() == alvo), None
        )
        if nome_real is None:
            raise KeyError(
                f"Aba {modulo.aba!r} não encontrada em {Path(caminho).name}. "
                f"Abas: {livro.sheetnames}"
            )
        aba = livro[nome_real]
        linhas = list(
            aba.iter_rows(min_row=modulo.linha_cabecalho + 1, values_only=True)
        )
    finally:
        livro.close()

    registros: list[Registro] = []

    for posicao, linha in enumerate(linhas):
        item = _texto(_indice(linha, modulo.col_item))
        if not item:
            continue

        telefone, motivo = normalizar_com_motivo(_indice(linha, modulo.col_telefone))
        empresa = _texto(_indice(linha, modulo.col_empresa)) or "(sem empresa)"
        numero_linha = modulo.linha_cabecalho + 1 + posicao

        # Rótulo do item: placa + frota, quando houver
        rotulo = item
        if modulo.col_frota is not None:
            frota = _texto(_indice(linha, modulo.col_frota))
            if frota:
                rotulo = f"{item} (frota {frota})"

        # Informações extras da linha (credencial, locador...)
        complementos = [
            f"{nome}: {_texto(_indice(linha, coluna))}"
            for coluna, nome in modulo.extras.items()
            if _texto(_indice(linha, coluna))
        ]

        alertas: list[tuple[str, dt.date | None, str]] = []

        for coluna, nome_doc in modulo.documentos.items():
            data = _para_data_frotas(_indice(linha, coluna))
            situacao = _situacao(data, hoje)
            if situacao:
                alertas.append((nome_doc, data, situacao))

        # CRLV é ano, não data: ano anterior ao vigente = vencido
        if modulo.col_crlv_ano is not None:
            ano = _texto(_indice(linha, modulo.col_crlv_ano))
            if ano.isdigit() and int(ano) < hoje.year:
                alertas.append((f"CRLV {ano}", None, "Vencido"))

        for nome_doc, data, situacao in alertas:
            registros.append(
                Registro(
                    item=rotulo,
                    validade=data,
                    status=situacao,
                    empresa=empresa,
                    telefone=telefone,
                    telefone_motivo=motivo,
                    telefone_exibicao=formatar_exibicao(telefone),
                    email="",
                    linha_planilha=numero_linha,
                    documento=(
                        nome_doc
                        if not complementos
                        else f"{nome_doc} ({'; '.join(complementos)})"
                    ),
                )
            )

    return registros


def _indice(linha: tuple, posicao: int | None):
    if posicao is None or posicao >= len(linha):
        return None
    return linha[posicao]


# ======================================================================
# MODULO: mensagens
# ======================================================================
"""Textos das mensagens de cobrança.

EDITE ESTE ARQUIVO para mudar o texto — não é preciso mexer no resto do código.

Como funciona:
  - {saudacao}  -> "Prezados(a)," (ou o que você colocar em SAUDACAO)
  - {itens}     -> lista dos documentos vencidos, montada automaticamente
  - {empresa}   -> nome da empresa
Qualquer um deles pode ser omitido do texto se não fizer sentido.
"""



ASSINATURA = "\n\nAtenciosamente,\nGestão de Frotas - Autoport"

TEMPLATES: dict[str, str] = {
    # ------------------------------------------------------------------
    "cavalo": (
        "Prezados(a), informo que a permissão portuária do(s) Cavalo(s) abaixo "
        "estão próximos do vencimento, gentileza enviar o CRLV do ano vigente e "
        "a ANTT-RNTRC para que seja renovado o permisso."
        "\n\n{itens}"
    ),
    # ------------------------------------------------------------------
    "motorista": (
        "Prezados(a), informo que a sua permissão portuária está próximo do "
        "vencimento, gentileza enviar sua Carteira de Trabalho Digital extraída "
        "do aplicativo juntamente com comprovante de residência em seu nome, "
        "ambos com data recente."
        "\n\n{itens}"
    ),
    # ------------------------------------------------------------------
    # RASCUNHO - texto não foi definido ainda. Substitua pelo texto oficial.
    "carreta": (
        "Prezados(a), informo que a permissão portuária da(s) carreta(s) abaixo "
        "está próxima do vencimento. Gentileza enviar a documentação atualizada "
        "para que seja renovado o permisso."
        "\n\n{itens}"
    ),
    # ------------------------------------------------------------------
    # RASCUNHO - texto não foi definido ainda. Substitua pelo texto oficial.
    "empresas": (
        "Prezados(a), informo que o cadastro portuário da empresa abaixo está "
        "próximo do vencimento. Gentileza enviar a documentação atualizada para "
        "renovação."
        "\n\n{itens}"
    ),
    # ------------------------------------------------------------------
    # ---------------------- PLANILHA MAPA FROTAS ----------------------
    # ------------------------------------------------------------------
    "cavalo_carreta": (
        "Prezados(a), Segue alerta de documentos vencidos e a vencer. Evite o "
        "bloqueio automático das frotas/motoristas no sistema, enviando as "
        "atualizações com antecedências."
        "\n\n{itens}"
    ),
    # ------------------------------------------------------------------
    "motoristas_agregados": (
        "Prezados(a), Segue alerta de documentos vencidos e a vencer. Evite o "
        "bloqueio automático das frotas/motoristas no sistema, enviando as "
        "atualizações com antecedências."
        "\n\n{itens}"
    ),
    # ------------------------------------------------------------------
    "contrato_aluguel": (
        "Prezados(a), Informo que o Locação da credencial está próxima do "
        "vencimento, segue dados abaixo para verificação."
        "\n\n{itens}"
    ),
}

# Módulos cujo texto ainda é rascunho — a interface avisa antes do envio.
RASCUNHOS = {"carreta", "empresas"}


def montar_itens(contato: Contato, rotulo_item: str) -> str:
    """Monta a lista de itens vencidos do contato."""
    linhas = []
    for reg in sorted(contato.registros, key=lambda r: (r.item, r.documento)):
        marcador = "[VENCIDO]" if reg.status == "Vencido" else "[a vencer]"
        if reg.validade is None:
            validade = ""
        else:
            validade = f" — validade {reg.validade_texto}"
        linhas.append(f"- {reg.descricao}{validade} {marcador}")
    return "\n".join(linhas)


def montar_mensagem(
    chave_modulo: str, contato: Contato, rotulo_item: str = "Item"
) -> str:
    """Devolve o texto final que será enviado a este contato."""
    template = TEMPLATES.get(chave_modulo)
    if template is None:
        raise KeyError(f"Não há template para o módulo {chave_modulo!r}")

    texto = template.format(
        itens=montar_itens(contato, rotulo_item),
        empresa=contato.empresa,
    )
    return texto + ASSINATURA


# ======================================================================
# MODULO: modulos
# ======================================================================

# alias: no arquivo unico os dois dicionarios convivem no mesmo escopo
MODULOS_PORTO = MODULOS
"""Registro único dos módulos das duas planilhas.

O app não precisa saber de qual arquivo veio cada card — pergunta aqui.
"""




# Nome amigável de cada planilha
FONTES = {
    "porto": "Mapa Porto (GERAL-NOVO_MAPA_PORTO.xlsx)",
    "frotas": "Mapa Frotas (Mapa_Frotas.xlsx)",
}


def todos_modulos() -> dict[str, object]:
    """chave -> Modulo | ModuloFrotas, na ordem de exibição."""
    combinado: dict[str, object] = {}
    combinado.update(MODULOS_PORTO)
    combinado.update(MODULOS_FROTAS)
    return combinado


def fonte_do_modulo(chave: str) -> str:
    return "frotas" if chave in MODULOS_FROTAS else "porto"


def rotulo_do_modulo(chave: str) -> str:
    return todos_modulos()[chave].rotulo  # type: ignore[attr-defined]


def rotulo_item(chave: str) -> str:
    return getattr(todos_modulos()[chave], "rotulo_item", "Item")


def ler(chave: str, caminho: Path | str, hoje: dt.date | None = None) -> list[Registro]:
    """Lê os registros pendentes de um módulo, seja qual for a planilha."""
    if chave in MODULOS_FROTAS:
        return ler_registros_frotas(caminho, MODULOS_FROTAS[chave], hoje)
    if chave in MODULOS_PORTO:
        return ler_registros(caminho, MODULOS_PORTO[chave], hoje)
    raise KeyError(f"Módulo desconhecido: {chave!r}")


def resumo_geral(
    caminhos: dict[str, str | Path | None], hoje: dt.date | None = None
) -> dict[str, dict]:
    """Contagens de todos os módulos. caminhos = {'porto': ..., 'frotas': ...}."""
    saida: dict[str, dict] = {}

    for chave in todos_modulos():
        fonte = fonte_do_modulo(chave)
        caminho = caminhos.get(fonte)

        vazio = {
            "fonte": fonte,
            "pendentes": 0,
            "vencidos": 0,
            "a_vencer": 0,
            "sem_telefone": 0,
            "contatos": 0,
        }

        if not caminho or not Path(caminho).exists():
            saida[chave] = {**vazio, "erro": "planilha não carregada"}
            continue

        try:
            registros = ler(chave, caminho, hoje)
        except (KeyError, ValueError, OSError) as erro:
            saida[chave] = {**vazio, "erro": str(erro)}
            continue

        saida[chave] = {
            **vazio,
            "erro": None,
            "pendentes": len(registros),
            "vencidos": sum(1 for r in registros if r.status == "Vencido"),
            "a_vencer": sum(
                1 for r in registros if r.status == "A Vencer em 30 dias"
            ),
            "sem_telefone": len(pendencias_de_cadastro(registros)),
            "contatos": len(agrupar_por_contato(registros)),
        }

    return saida


# ======================================================================
# MODULO: historico
# ======================================================================
"""Registro de envios e bloqueio de reenvio dentro de 5 dias."""



CAMINHO_PADRAO = Path("historico_envios.csv")
INTERVALO_DIAS = 5

CABECALHO = [
    "data_hora",
    "modulo",
    "empresa",
    "telefone",
    "qtd_itens",
    "itens",
    "status",
    "erro",
]

hist_ENVIADO = "enviado"
hist_SIMULADO = "simulado"
hist_FALHOU = "falhou"


def hist_garantir_arquivo(caminho: Path | str = CAMINHO_PADRAO) -> Path:
    caminho = Path(caminho)
    if not caminho.exists():
        caminho.parent.mkdir(parents=True, exist_ok=True)
        with caminho.open("w", newline="", encoding="utf-8") as arq:
            csv.writer(arq).writerow(CABECALHO)
    return caminho


def hist_registrar(
    modulo: str,
    empresa: str,
    telefone: str,
    itens: list[str],
    status: str,
    erro: str = "",
    caminho: Path | str = CAMINHO_PADRAO,
) -> None:
    caminho = hist_garantir_arquivo(caminho)
    with Path(caminho).open("a", newline="", encoding="utf-8") as arq:
        csv.writer(arq).writerow(
            [
                dt.datetime.now().isoformat(timespec="seconds"),
                modulo,
                empresa,
                telefone,
                len(itens),
                " | ".join(itens),
                status,
                erro,
            ]
        )


def hist_ultimo_envio(
    modulo: str, telefone: str, caminho: Path | str = CAMINHO_PADRAO
) -> dt.datetime | None:
    """Data do último envio REAL (simulações não contam)."""
    caminho = Path(caminho)
    if not caminho.exists():
        return None

    ultima: dt.datetime | None = None
    with caminho.open(newline="", encoding="utf-8") as arq:
        for linha in csv.DictReader(arq):
            if linha.get("modulo") != modulo:
                continue
            if linha.get("telefone") != telefone:
                continue
            if linha.get("status") != ENVIADO:
                continue
            try:
                quando = dt.datetime.fromisoformat(linha["data_hora"])
            except (ValueError, KeyError):
                continue
            if ultima is None or quando > ultima:
                ultima = quando
    return ultima


def hist_dias_desde_ultimo(
    modulo: str, telefone: str, caminho: Path | str = CAMINHO_PADRAO
) -> int | None:
    ultima = hist_ultimo_envio(modulo, telefone, caminho)
    if ultima is None:
        return None
    return (dt.datetime.now() - ultima).days


def hist_pode_enviar(
    modulo: str,
    telefone: str,
    caminho: Path | str = CAMINHO_PADRAO,
    intervalo: int = INTERVALO_DIAS,
) -> tuple[bool, str]:
    """(pode, motivo). Motivo em texto para mostrar na interface."""
    dias = hist_dias_desde_ultimo(modulo, telefone, caminho)
    if dias is None:
        return True, "nunca cobrado"
    if dias >= intervalo:
        return True, f"último envio há {dias} dia(s)"
    faltam = intervalo - dias
    return False, f"cobrado há {dias} dia(s) — liberado em {faltam} dia(s)"


def hist_carregar_historico(caminho: Path | str = CAMINHO_PADRAO) -> list[dict]:
    caminho = Path(caminho)
    if not caminho.exists():
        return []
    with caminho.open(newline="", encoding="utf-8") as arq:
        return list(csv.DictReader(arq))


# ======================================================================
# MODULO: whatsapp
# ======================================================================
"""Envio via WhatsApp Web.

IMPORTANTE: automação do WhatsApp Web viola os termos de uso do WhatsApp e pode
resultar em bloqueio do número. O modo simulação é o padrão justamente para você
validar tudo antes de arriscar o número da empresa.

A alternativa sem risco de bloqueio é a WhatsApp Business API oficial (Meta,
Twilio, Zenvia) — tem custo por mensagem e exige templates aprovados.
"""



PAUSA_MINIMA = 20
PAUSA_MAXIMA = 40
LIMITE_PADRAO = 50


@dataclass
class ResultadoEnvio:
    telefone: str
    sucesso: bool
    simulado: bool
    erro: str = ""


class EnviadorBase:
    """Interface comum aos enviadores."""

    def enviar(self, telefone: str, mensagem: str) -> ResultadoEnvio:
        raise NotImplementedError

    def encerrar(self) -> None:
        pass


class EnviadorSimulado(EnviadorBase):
    """Não envia nada. Só devolve sucesso, para validar o fluxo."""

    def enviar(self, telefone: str, mensagem: str) -> ResultadoEnvio:
        time.sleep(0.15)  # dá sensação de progresso na barra
        return ResultadoEnvio(telefone=telefone, sucesso=True, simulado=True)


class EnviadorNavegador(EnviadorBase):
    """Abre wa.me no navegador. Exige um clique manual em cada mensagem.

    É o modo mais seguro: o WhatsApp não vê automação de digitação, e você
    confere cada mensagem antes de mandar. Serve bem para os volumes daqui
    (7 a 14 mensagens por módulo).
    """

    def enviar(self, telefone: str, mensagem: str) -> ResultadoEnvio:
        try:
            url = (
                f"https://web.whatsapp.com/send?phone={telefone}"
                f"&text={urllib.parse.quote(mensagem)}"
            )
            webbrowser.open(url)
            return ResultadoEnvio(telefone=telefone, sucesso=True, simulado=False)
        except Exception as erro:  # noqa: BLE001
            return ResultadoEnvio(
                telefone=telefone, sucesso=False, simulado=False, erro=str(erro)
            )


class EnviadorAutomatico(EnviadorBase):
    """Envio sem clique, via pywhatkit. É o modo com maior risco de bloqueio."""

    def __init__(self) -> None:
        try:
            import pywhatkit  # noqa: F401
        except ImportError as erro:
            raise RuntimeError(
                "pywhatkit não está instalado. Rode: pip install pywhatkit"
            ) from erro
        self._kit = __import__("pywhatkit")

    def enviar(self, telefone: str, mensagem: str) -> ResultadoEnvio:
        try:
            self._kit.sendwhatmsg_instantly(
                phone_no=f"+{telefone}",
                message=mensagem,
                wait_time=15,
                tab_close=True,
                close_time=3,
            )
            return ResultadoEnvio(telefone=telefone, sucesso=True, simulado=False)
        except Exception as erro:  # noqa: BLE001
            return ResultadoEnvio(
                telefone=telefone, sucesso=False, simulado=False, erro=str(erro)
            )


def link_whatsapp(telefone: str, mensagem: str) -> str:
    """Link que abre o WhatsApp Web NO NAVEGADOR DE QUEM CLICA.

    É o único modo que funciona com o app hospedado (Streamlit Cloud, etc.):
    o servidor não abre nada, só entrega o link pronto. Quem clica é o usuário,
    na máquina dele, com a sessão dele do WhatsApp.
    """
    return (
        f"https://wa.me/{telefone}?text={urllib.parse.quote(mensagem)}"
    )


def criar_enviador(modo: str) -> EnviadorBase:
    """modo: 'simulacao' | 'navegador' | 'automatico'."""
    if modo == "simulacao":
        return EnviadorSimulado()
    if modo == "navegador":
        return EnviadorNavegador()
    if modo == "automatico":
        return EnviadorAutomatico()
    raise ValueError(f"Modo de envio desconhecido: {modo!r}")


def pausa_aleatoria(minimo: int = PAUSA_MINIMA, maximo: int = PAUSA_MAXIMA) -> float:
    """Intervalo entre mensagens. Rajada é o que mais dispara bloqueio."""
    return random.uniform(minimo, maximo)


"""Automações Gestão de Frotas - Autoport.

Rodar com:  streamlit run app.py
"""



import streamlit as st


RAIZ = Path(__file__).parent
PLANILHAS_PADRAO = {
    "porto": RAIZ / "dados" / "GERAL-NOVO_MAPA_PORTO.xlsx",
    "frotas": RAIZ / "dados" / "Mapa_Frotas.xlsx",
}
LOGO = RAIZ / "assets" / "logo_autoport.png"

PRETO = "#0D0D0D"
VERMELHO = "#8B0000"
VERMELHO_CLARO = "#B22222"
CINZA_ESCURO = "#2B2B2B"
CINZA_CLARO = "#B0B0B0"

st.set_page_config(
    page_title="Automações Gestão de Frotas - Autoport",
    page_icon="🚛",
    layout="wide",
)

CSS = f"""
<style>
    .stApp {{ background-color: {PRETO}; }}

    .cabecalho {{
        border-bottom: 3px solid {VERMELHO};
        padding-bottom: 14px;
        margin-bottom: 28px;
    }}
    .titulo {{
        font-size: 30px; font-weight: 700; color: #FFFFFF;
        letter-spacing: .5px; margin: 0;
    }}
    .subtitulo {{ color: {CINZA_CLARO}; font-size: 14px; margin-top: 4px; }}

    .cartao {{
        background: {CINZA_ESCURO};
        border-left: 5px solid {VERMELHO};
        border-radius: 6px;
        padding: 18px 20px;
        margin-bottom: 10px;
        min-height: 148px;
    }}
    .cartao-titulo {{
        color: #FFFFFF; font-size: 15px; font-weight: 700;
        text-transform: uppercase; letter-spacing: .6px; margin-bottom: 10px;
    }}
    .cartao-numero {{
        color: {VERMELHO_CLARO}; font-size: 40px; font-weight: 800;
        line-height: 1;
    }}
    .cartao-legenda {{ color: {CINZA_CLARO}; font-size: 12px; margin-top: 6px; }}
    .cartao-alerta {{ color: #E8A33D; font-size: 12px; margin-top: 4px; }}

    .faixa {{
        background: {CINZA_ESCURO}; border-left: 4px solid {VERMELHO};
        padding: 12px 16px; border-radius: 4px; margin-bottom: 16px;
    }}
    .previa {{
        background: #111B14; border: 1px solid #2F4F3A; border-radius: 6px;
        padding: 16px; white-space: pre-wrap; color: #E8E8E8;
        font-family: -apple-system, "Segoe UI", sans-serif; font-size: 13px;
        line-height: 1.55;
    }}

    div.stButton > button {{
        background: {VERMELHO}; color: #FFFFFF; border: none;
        border-radius: 4px; font-weight: 600;
    }}
    div.stButton > button:hover {{ background: {VERMELHO_CLARO}; color: #FFF; }}

    section[data-testid="stSidebar"] {{ background: #141414; }}
    #MainMenu, footer {{ visibility: hidden; }}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)


# ---------------------------------------------------------------- estado
def estado(chave: str, padrao):
    if chave not in st.session_state:
        st.session_state[chave] = padrao
    return st.session_state[chave]


estado("modulo_aberto", None)
estado("selecionados", {})


# ---------------------------------------------------------------- dados
@st.cache_data(show_spinner="Lendo as planilhas...")
def carregar_resumo(caminhos: dict, quando: str):
    return resumo_geral(caminhos, dt.date.fromisoformat(quando))


@st.cache_data(show_spinner="Carregando módulo...")
def carregar_modulo(caminho: str, chave: str, quando: str):
    registros = ler(chave, caminho, dt.date.fromisoformat(quando))
    return registros, agrupar_por_contato(registros), pendencias_de_cadastro(registros)


# ---------------------------------------------------------------- topo
def cabecalho():
    col_logo, col_titulo = st.columns([1, 6])
    with col_logo:
        if LOGO.exists():
            st.image(str(LOGO), width=110)
        else:
            st.markdown(
                f"<div style='background:{VERMELHO};color:#fff;padding:18px 10px;"
                "border-radius:6px;text-align:center;font-weight:800;"
                "letter-spacing:1px;font-size:15px'>AUTOPORT</div>",
                unsafe_allow_html=True,
            )
    with col_titulo:
        st.markdown(
            "<div class='cabecalho'>"
            "<p class='titulo'>Automações Gestão de Frotas - Autoport</p>"
            "<p class='subtitulo'>Cobrança de documentação portuária via WhatsApp</p>"
            "</div>",
            unsafe_allow_html=True,
        )


# ---------------------------------------------------------------- sidebar
def barra_lateral():
    st.sidebar.markdown("### Planilhas")

    caminhos: dict[str, str | None] = {}

    for fonte, descricao in FONTES.items():
        enviado = st.sidebar.file_uploader(
            descricao, type=["xlsx"], key=f"up_{fonte}"
        )
        if enviado is not None:
            destino = RAIZ / "dados" / f"carregada_{fonte}.xlsx"
            destino.parent.mkdir(exist_ok=True)
            destino.write_bytes(enviado.getbuffer())
            caminhos[fonte] = str(destino)
            st.sidebar.success(f"{fonte}: arquivo carregado")
        else:
            padrao = PLANILHAS_PADRAO[fonte]
            caminhos[fonte] = str(padrao) if padrao.exists() else None
            if caminhos[fonte] is None:
                st.sidebar.warning(f"{fonte}: nenhuma planilha")

    if not any(caminhos.values()):
        st.sidebar.info("Envie ao menos uma planilha para começar.")
        st.markdown(
            "<div class='faixa'><b>Comece enviando as planilhas.</b><br>"
            "Use os campos da barra lateral à esquerda. Você pode enviar as "
            "duas ou só uma delas — os cards da planilha que faltar ficam "
            "indisponíveis.</div>",
            unsafe_allow_html=True,
        )
        st.stop()

    referencia = st.sidebar.date_input(
        "Data de referência",
        value=dt.date.today(),
        help="Base para calcular vencido e a vencer em 30 dias.",
    )

    st.sidebar.markdown("---")
    st.sidebar.markdown("### Modo de envio")

    modo_rotulo = st.sidebar.radio(
        "Como enviar",
        [
            "Links do WhatsApp (recomendado)",
            "Simulação (não envia nada)",
            "Abrir no servidor",
            "Envio automático",
        ],
        index=0,
        label_visibility="collapsed",
    )
    modo = {
        "Links do WhatsApp (recomendado)": "links",
        "Simulação (não envia nada)": "simulacao",
        "Abrir no servidor": "navegador",
        "Envio automático": "automatico",
    }[modo_rotulo]

    if modo == "links":
        st.sidebar.info(
            "Gera um botão por contato. O WhatsApp Web abre **na sua máquina**, "
            "com a mensagem pronta. Único modo que funciona com o app hospedado."
        )
    elif modo == "simulacao":
        st.sidebar.info("Nada será enviado. Use para conferir os textos.")
    elif modo == "navegador":
        st.sidebar.warning(
            "Abre o navegador **do servidor**. Só funciona se o app estiver "
            "rodando na sua própria máquina."
        )
    else:
        st.sidebar.error(
            "Envio sem confirmação, via pywhatkit. Exige o app rodando local "
            "com o WhatsApp Web logado. Viola os termos de uso do WhatsApp."
        )

    limite = st.sidebar.number_input(
        "Máximo por execução", min_value=1, max_value=300, value=LIMITE_PADRAO
    )

    st.sidebar.markdown("---")
    if st.sidebar.button("Recarregar planilhas"):
        st.cache_data.clear()
        st.rerun()

    return caminhos, referencia.isoformat(), modo, int(limite)


# ---------------------------------------------------------------- cards
def tela_inicial(caminhos: dict, quando: str):
    dados = carregar_resumo(caminhos, quando)

    for fonte, descricao in FONTES.items():
        chaves = [c for c in todos_modulos() if fonte_do_modulo(c) == fonte]
        if not chaves:
            continue

        st.markdown(f"#### {descricao}")
        colunas = st.columns(max(len(chaves), 1))

        for coluna, chave in zip(colunas, chaves):
            info = dados[chave]
            with coluna:
                _cartao(chave, info)

        st.markdown("")

    validos = [v for v in dados.values() if not v["erro"]]
    total = sum(v["contatos"] for v in validos)
    sem_tel = sum(v["sem_telefone"] for v in validos)
    st.markdown(
        f"<div class='faixa'>Total geral: <b>{total}</b> mensagem(ns) a disparar. "
        f"<span style='color:{CINZA_CLARO}'>{sem_tel} pendência(s) sem telefone "
        "válido não entram na fila.</span></div>",
        unsafe_allow_html=True,
    )


def _cartao(chave: str, info: dict):
    rotulo = rotulo_do_modulo(chave)

    if info["erro"]:
        st.markdown(
            f"<div class='cartao'><div class='cartao-titulo'>{rotulo}</div>"
            f"<div style='color:#E06C6C;font-size:12px'>{info['erro'][:130]}</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        return

    alerta = ""
    if info["sem_telefone"]:
        alerta = (
            f"<div class='cartao-alerta'>{info['sem_telefone']} sem telefone "
            "válido</div>"
        )

    st.markdown(
        f"""<div class='cartao'>
            <div class='cartao-titulo'>{rotulo}</div>
            <div class='cartao-numero'>{info['contatos']}</div>
            <div class='cartao-legenda'>
                contato(s) a cobrar<br>
                {info['pendentes']} alerta(s):
                {info['vencidos']} vencido(s),
                {info['a_vencer']} a vencer
            </div>{alerta}
        </div>""",
        unsafe_allow_html=True,
    )
    if st.button(f"Abrir {rotulo}", key=f"btn_{chave}", use_container_width=True):
        st.session_state["modulo_aberto"] = chave
        st.rerun()


# ---------------------------------------------------------------- módulo
def tela_modulo(chave: str, caminhos: dict, quando: str, modo: str, limite: int):
    fonte = fonte_do_modulo(chave)
    caminho = caminhos.get(fonte)

    if st.button("← Voltar aos módulos"):
        st.session_state["modulo_aberto"] = None
        st.rerun()

    st.markdown(f"## {rotulo_do_modulo(chave)}")
    st.caption(FONTES[fonte])

    if not caminho:
        st.error("A planilha deste módulo não foi carregada.")
        return

    if chave in RASCUNHOS:
        st.warning(
            "O texto deste módulo ainda é um **rascunho** — o texto oficial não "
            "foi definido. Edite `core/mensagens.py` antes de enviar de verdade."
        )

    try:
        registros, contatos, sem_telefone = carregar_modulo(caminho, chave, quando)
    except (KeyError, ValueError, OSError) as erro:
        st.error(f"Não consegui ler a planilha: {erro}")
        return

    if not registros:
        st.success("Nenhuma pendência neste módulo.")
        return

    filtro = st.radio(
        "Situação",
        ["Todos", "Somente vencidos", "Somente a vencer"],
        horizontal=True,
    )

    def passa(contato: Contato) -> bool:
        if filtro == "Somente vencidos":
            return any(r.status == "Vencido" for r in contato.registros)
        if filtro == "Somente a vencer":
            return any(r.status == "A Vencer em 30 dias" for r in contato.registros)
        return True

    visiveis = [c for c in contatos if passa(c)]

    aba_fila, aba_pendencias, aba_historico = st.tabs(
        [f"Fila de envio ({len(visiveis)})",
         f"Sem telefone ({len(sem_telefone)})",
         "Histórico"]
    )

    with aba_fila:
        _aba_fila(chave, visiveis, modo, limite)

    with aba_pendencias:
        _aba_sem_telefone(sem_telefone)

    with aba_historico:
        _aba_historico(chave)


def _aba_fila(chave, contatos, modo, limite):
    if not contatos:
        st.info("Nenhum contato neste filtro.")
        return

    col_a, col_b = st.columns([1, 1])
    with col_a:
        if st.button("Selecionar todos", use_container_width=True):
            for c in contatos:
                pode, _ = hist_pode_enviar(chave, c.telefone)
                st.session_state["selecionados"][f"{chave}:{c.telefone}"] = pode
            st.rerun()
    with col_b:
        if st.button("Limpar seleção", use_container_width=True):
            for c in contatos:
                st.session_state["selecionados"][f"{chave}:{c.telefone}"] = False
            st.rerun()

    st.markdown("---")

    escolhidos: list[Contato] = []

    for contato in contatos:
        pode, motivo = hist_pode_enviar(chave, contato.telefone)
        marca = f"{chave}:{contato.telefone}"
        padrao = st.session_state["selecionados"].get(marca, pode)

        col_check, col_dados = st.columns([1, 11])

        with col_check:
            marcado = st.checkbox(
                "sel", key=f"chk_{marca}", value=padrao, label_visibility="collapsed"
            )
            st.session_state["selecionados"][marca] = marcado

        with col_dados:
            titulo = f"**{contato.empresa}** — {contato.telefone_exibicao}"
            etiquetas = []
            if not pode:
                etiquetas.append(f":orange[{motivo}]")
            if contato.tem_numero_inferido:
                etiquetas.append(":orange[número com 9 inferido — confira]")
            sufixo = "  ·  ".join(etiquetas)

            with st.expander(
                f"{titulo}  —  {contato.qtd} item(ns)  {sufixo}", expanded=False
            ):
                st.dataframe(
                    pd.DataFrame(
                        [
                            {
                                rotulo_item(chave): r.item,
                                "Documento": r.documento or "—",
                                "Validade": r.validade_texto,
                                "Situação": r.status,
                                "Linha": r.linha_planilha,
                            }
                            for r in contato.registros
                        ]
                    ),
                    hide_index=True,
                    use_container_width=True,
                )
                st.markdown("**Mensagem que será enviada:**")
                st.markdown(
                    f"<div class='previa'>{montar_mensagem(chave, contato)}</div>",
                    unsafe_allow_html=True,
                )

        if marcado:
            escolhidos.append(contato)

    st.markdown("---")

    if not escolhidos:
        st.info("Selecione ao menos um contato.")
        return

    bloqueados = [
        c for c in escolhidos if not hist_pode_enviar(chave, c.telefone)[0]
    ]

    st.markdown(f"**{len(escolhidos)}** contato(s) selecionado(s).")

    if bloqueados:
        st.warning(
            f"{len(bloqueados)} contato(s) foram cobrados há menos de 5 dias. "
            "Marque a opção abaixo para cobrar mesmo assim."
        )
        forcar = st.checkbox("Ignorar o intervalo de 5 dias")
        if not forcar:
            escolhidos = [c for c in escolhidos if c not in bloqueados]

    if len(escolhidos) > limite:
        st.warning(f"Limite de {limite} por execução. Serão enviados os {limite} primeiros.")
        escolhidos = escolhidos[:limite]

    if modo == "links":
        if st.button(
            f"Gerar {len(escolhidos)} link(s) do WhatsApp",
            type="primary",
            use_container_width=True,
        ):
            st.session_state["links_gerados"] = {
                "modulo": chave,
                "contatos": [c.telefone for c in escolhidos],
            }
        if st.session_state.get("links_gerados", {}).get("modulo") == chave:
            _painel_links(chave, escolhidos)
        return

    rotulo = {
        "simulacao": f"Simular {len(escolhidos)} envio(s)",
        "navegador": f"Abrir {len(escolhidos)} conversa(s) no WhatsApp",
        "automatico": f"ENVIAR {len(escolhidos)} mensagem(ns) agora",
    }[modo]

    if st.button(rotulo, type="primary", use_container_width=True):
        _disparar(chave, escolhidos, modo)


def _painel_links(chave, contatos):
    """Lista de botões wa.me. O clique acontece na máquina do usuário."""
    st.markdown("---")
    st.markdown("### Envio por link")
    st.markdown(
        "Clique em **Abrir WhatsApp** de cada linha. A conversa abre em uma aba "
        "nova, com a mensagem já escrita — você só confere e aperta enviar. "
        "Depois marque **Enviado** para registrar no histórico e ativar o "
        "bloqueio de 5 dias."
    )

    for contato in contatos:
        texto = montar_mensagem(chave, contato)
        url = link_whatsapp(contato.telefone, texto)
        marca = f"{chave}:{contato.telefone}"

        col_nome, col_link, col_ok = st.columns([6, 2, 2])

        with col_nome:
            st.markdown(
                f"**{contato.empresa}** · {contato.telefone_exibicao} · "
                f"{contato.qtd} item(ns)"
            )
        with col_link:
            st.link_button("Abrir WhatsApp", url, use_container_width=True)
        with col_ok:
            ja = st.session_state.get(f"env_{marca}", False)
            if ja:
                st.markdown(
                    "<div style='color:#5FBF6A;text-align:center;padding-top:6px'>"
                    "registrado</div>",
                    unsafe_allow_html=True,
                )
            elif st.button("Enviado", key=f"ok_{marca}", use_container_width=True):
                hist_registrar(
                    modulo=chave,
                    empresa=contato.empresa,
                    telefone=contato.telefone,
                    itens=[r.item for r in contato.registros],
                    status=hist_ENVIADO,
                )
                st.session_state[f"env_{marca}"] = True
                st.rerun()

    st.markdown("---")
    linhas = [
        {
            "Empresa": c.empresa,
            "Telefone": c.telefone_exibicao,
            "Itens": c.qtd,
            "Link": link_whatsapp(c.telefone, montar_mensagem(chave, c)),
            "Mensagem": montar_mensagem(chave, c),
        }
        for c in contatos
    ]
    st.download_button(
        "Baixar links e mensagens (CSV)",
        pd.DataFrame(linhas).to_csv(index=False).encode("utf-8-sig"),
        file_name=f"links_whatsapp_{chave}.csv",
        mime="text/csv",
    )


def _disparar(chave, contatos, modo):
    try:
        enviador = criar_enviador(modo)
    except RuntimeError as erro:
        st.error(str(erro))
        return

    barra = st.progress(0.0)
    area = st.empty()
    linhas = []

    for indice, contato in enumerate(contatos, start=1):
        area.markdown(
            f"Processando **{contato.empresa}** "
            f"({contato.telefone_exibicao}) — {indice}/{len(contatos)}"
        )

        texto = montar_mensagem(chave, contato)
        resultado = enviador.enviar(contato.telefone, texto)
        itens = [r.item for r in contato.registros]

        if resultado.simulado:
            situacao = hist_SIMULADO
        elif resultado.sucesso:
            situacao = hist_ENVIADO
        else:
            situacao = hist_FALHOU

        hist_registrar(
            modulo=chave,
            empresa=contato.empresa,
            telefone=contato.telefone,
            itens=itens,
            status=situacao,
            erro=resultado.erro,
        )

        linhas.append(
            {
                "Empresa": contato.empresa,
                "Telefone": contato.telefone_exibicao,
                "Itens": len(itens),
                "Resultado": situacao,
                "Erro": resultado.erro or "",
            }
        )

        barra.progress(indice / len(contatos))

        if indice < len(contatos) and modo != "simulacao":
            espera = pausa_aleatoria()
            area.markdown(f"Aguardando {espera:.0f}s antes do próximo envio...")
            time.sleep(espera)

    enviador.encerrar()
    area.empty()
    barra.empty()

    st.success(f"Processo concluído — {len(contatos)} contato(s).")
    st.dataframe(pd.DataFrame(linhas), hide_index=True, use_container_width=True)

    if modo == "simulacao":
        st.info(
            "Isto foi uma **simulação**. Nenhuma mensagem saiu e o intervalo de "
            "5 dias não foi consumido."
        )


def _aba_sem_telefone(registros):
    if not registros:
        st.success("Todos os pendentes têm telefone cadastrado.")
        return

    st.markdown(
        "Estas pendências **não serão cobradas** por falta de telefone válido. "
        "Corrija na planilha e recarregue."
    )
    tabela = pd.DataFrame(
        [
            {
                "Item": r.item,
                "Documento": r.documento or "—",
                "Empresa": r.empresa,
                "Validade": r.validade_texto,
                "Situação": r.status,
                "Motivo": (
                    "sem cadastro"
                    if r.telefone_motivo == "sem_cadastro"
                    else "número inválido"
                ),
                "Linha": r.linha_planilha,
            }
            for r in registros
        ]
    )
    st.dataframe(tabela, hide_index=True, use_container_width=True)
    st.download_button(
        "Baixar lista (CSV)",
        tabela.to_csv(index=False).encode("utf-8-sig"),
        file_name="pendencias_sem_telefone.csv",
        mime="text/csv",
    )


def _aba_historico(chave):
    registros = [
        linha for linha in hist_carregar_historico() if linha.get("modulo") == chave
    ]
    if not registros:
        st.info("Nenhum envio registrado ainda para este módulo.")
        return
    tabela = pd.DataFrame(registros).sort_values("data_hora", ascending=False)
    st.dataframe(tabela, hide_index=True, use_container_width=True)


# ---------------------------------------------------------------- main
def main():
    cabecalho()
    caminhos, quando, modo, limite = barra_lateral()

    aberto = st.session_state["modulo_aberto"]
    if aberto is None:
        tela_inicial(caminhos, quando)
    else:
        tela_modulo(aberto, caminhos, quando, modo, limite)


if __name__ == "__main__":
    main()
